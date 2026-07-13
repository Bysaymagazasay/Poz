#!/usr/bin/env python3
from __future__ import annotations

import csv
import importlib.util
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
BASE_PATH = ROOT / "scripts" / "build_mechanical_history.py"
OUT_DIR = ROOT / "generated" / "mechanical-history-last10"
MIN_YEAR = 2017
MAX_YEAR = 2026

spec = importlib.util.spec_from_file_location("mechanical_history", BASE_PATH)
base = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(base)


def year_of(period: str) -> int:
    import re
    m = re.search(r"(19\d{2}|20\d{2})", str(period or ""))
    return int(m.group(1)) if m else 0


def latest_master(history, mappings):
    old_to_new = {}
    for m in mappings:
        old = base.normalize_code(m.get("old_poz_no", ""))
        new = base.normalize_code(m.get("new_poz_no", ""))
        if old and new:
            old_to_new[old] = (new, m)

    by_code = defaultdict(list)
    for row in history:
        by_code[base.normalize_code(row.poz_no)].append(row)
    for rows in by_code.values():
        rows.sort(key=lambda r: (base.period_sort_key(r.period), r.source_page), reverse=True)

    result = []
    for code in sorted(by_code):
        rows = by_code[code]
        latest = rows[0]
        previous = rows[1] if len(rows) > 1 else None
        old_numbers = sorted(old for old, (new, _m) in old_to_new.items() if new == code and old in by_code)
        mapped_new = old_to_new.get(code, ("", {}))[0]
        status = "Güncel numara"
        if old_numbers:
            status = "Resmî eski numaraları var"
        elif mapped_new:
            status = "Resmî olarak yeni numaraya değişmiş"
        result.append({
            "Poz No": code,
            "Poz Tanımı": "",
            "Birim": latest.unit,
            "Son Yayın Fiyatı": latest.price,
            "Son Yayın Dönemi": latest.period,
            "Geçerlilik Tarihi": latest.validity_date,
            "Eski Poz Numaraları": " | ".join(old_numbers),
            "Yeni Poz Numarası": mapped_new,
            "Numara Durumu": status,
            "Önceki Yayın Dönemi": previous.period if previous else "",
            "Önceki Fiyat": previous.price if previous else "",
            "Yayın Sayısı": len(rows),
            "Kaynak": latest.source_url,
            "Kaynak Sayfa": latest.source_page,
            "Kontrol Notu": "Poz adı ikinci işlemde doldurulacak." if not mapped_new else "Eski/yeni numara resmî değişiklik belgesinden alınmıştır.",
        })
    return result


def write_xlsx(master, history, mappings, sources):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "CSB_Mekanik_Pozlar_Son_Yayin_Fiyatlari_2017_2026.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Son Yayın Listesi"

    headers = list(master[0].keys()) if master else ["Poz No", "Poz Tanımı", "Birim", "Son Yayın Fiyatı", "Son Yayın Dönemi"]
    ws.append(headers)
    for row in master:
        ws.append([row.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34
    widths = {"A":18,"B":42,"C":12,"D":18,"E":18,"F":16,"G":28,"H":22,"I":27,"J":20,"K":16,"L":12,"M":48,"N":13,"O":42}
    for col, width in widths.items(): ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        row[3].number_format = '#,##0.00'
        if len(row) > 10: row[10].number_format = '#,##0.00'
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    if ws.max_row > 1:
        tab = Table(displayName="MekanikSonYayin", ref=ws.dimensions)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws.add_table(tab)

    raw = wb.create_sheet("Tüm Yayın Kayıtları")
    raw_headers = ["Poz No","Birim","Fiyat","Dönem","Geçerlilik Tarihi","Kaynak","Sayfa"]
    raw.append(raw_headers)
    for r in sorted(history, key=lambda x:(base.period_sort_key(x.period), base.normalize_code(x.poz_no)), reverse=True):
        raw.append([r.poz_no,r.unit,r.price,r.period,r.validity_date,r.source_url,r.source_page])
    for c in raw[1]: c.fill=header_fill; c.font=header_font
    raw.freeze_panes="A2"; raw.auto_filter.ref=raw.dimensions
    for col,w in zip("ABCDEFG",[18,12,18,16,16,52,10]): raw.column_dimensions[col].width=w
    for cell in raw["C"][1:]: cell.number_format='#,##0.00'

    mp = wb.create_sheet("Eski Yeni Poz Eşleştirme")
    map_headers=["Eşleştirme Dönemi","Eski Poz No","Yeni Poz No","Kaynak","Kanıt","Eşleştirme Türü"]
    mp.append(map_headers)
    for m in mappings: mp.append([m.get("mapping_period",""),m.get("old_poz_no",""),m.get("new_poz_no",""),m.get("source_url",""),m.get("evidence",""),m.get("mapping_type","")])
    for c in mp[1]: c.fill=header_fill; c.font=header_font
    mp.freeze_panes="A2"; mp.auto_filter.ref=mp.dimensions
    for col,w in zip("ABCDEF",[18,20,20,52,70,18]): mp.column_dimensions[col].width=w

    src = wb.create_sheet("Kaynak Kontrolü")
    src_headers=["Dönem","Kaynak","Tür","Durum","Sayfa Sayısı","Mekanik Başlangıç","Mekanik Bitiş","Alınan Satır","Not"]
    src.append(src_headers)
    for s in sources: src.append([s.period,s.url,s.kind,s.status,s.pages,s.mechanical_start_page,s.mechanical_end_page,s.parsed_rows,s.note])
    for c in src[1]: c.fill=header_fill; c.font=header_font
    src.freeze_panes="A2"; src.auto_filter.ref=src.dimensions
    for col,w in zip("ABCDEFGHI",[16,55,16,25,14,18,16,14,60]): src.column_dimensions[col].width=w

    info = wb.create_sheet("Açıklama", 0)
    info.append(["ÇŞİDB / Yüksek Fen Kurulu Mekanik Pozları – Son Yayın Fiyatları"])
    info.append(["Kapsam", "2017–2026 arasındaki resmî İnşaat ve Tesisat Birim Fiyatları; ara dönem yayınları dâhil."])
    info.append(["Yöntem", "En yeni dönemden eskiye gidilerek her pozun son yayımlanan fiyatı seçilmiştir."])
    info.append(["Poz numarası değişiklikleri", "Yalnız resmî Değiştirilen Poz Numaraları belgelerindeki eşleştirmeler kullanılmıştır."])
    info.append(["Poz tanımları", "Bu aşamada özellikle boş bırakılmıştır; ikinci işlemde doldurulacaktır."])
    info.append(["Oluşturulma", datetime.now(timezone.utc).isoformat()])
    info.column_dimensions["A"].width=34; info.column_dimensions["B"].width=110
    info["A1"].font=Font(size=16,bold=True,color="17365D")
    info.merge_cells("A1:B1")
    for row in info.iter_rows():
        for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)
    wb.save(path)
    return path


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    sources, mapping_sources = base.collect_sources()
    sources = [s for s in sources if MIN_YEAR <= year_of(s.period) <= MAX_YEAR]
    mapping_sources = [s for s in mapping_sources if MIN_YEAR <= year_of(s.period) <= MAX_YEAR]
    sources.sort(key=lambda s: base.period_sort_key(s.period), reverse=True)

    history=[]; all_sources=[]
    for i,s in enumerate(sources,1):
        print(f"[{i}/{len(sources)}] {s.period} {s.url}", flush=True)
        try: history.extend(base.parse_price_book(s))
        except Exception as exc: s.status="error"; s.note=str(exc)[:1000]
        all_sources.append(s)
    mappings=[]
    for s in mapping_sources:
        try: mappings.extend(base.parse_code_map(s))
        except Exception as exc: s.status="error"; s.note=str(exc)[:1000]
        all_sources.append(s)

    dedup={}
    for r in history:
        key=(base.normalize_code(r.poz_no),r.period)
        if key not in dedup or r.source_page > dedup[key].source_page: dedup[key]=r
    history=list(dedup.values())
    master=latest_master(history,mappings)
    path=write_xlsx(master,history,mappings,all_sources)
    summary={"generated_at":datetime.now(timezone.utc).isoformat(),"range":"2017-2026","sources":len(sources),"raw_records":len(history),"master_records":len(master),"official_mappings":len(mappings),"xlsx":path.name}
    (OUT_DIR/"summary.json").write_text(json.dumps(summary,ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps(summary,ensure_ascii=False))
    return 0 if master else 2

if __name__ == "__main__":
    raise SystemExit(main())
