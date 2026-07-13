#!/usr/bin/env python3
from pathlib import Path
import csv, json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path('generated/mechanical-history-2017-2026')
MASTER = OUT / 'mechanical_master_latest.csv'
RAW = OUT / 'mechanical_history_raw.csv'
MAP = OUT / 'mechanical_code_map.csv'
SOURCES = OUT / 'mechanical_sources.csv'
XLSX = OUT / 'CSB_Mekanik_Pozlar_Son_Yayin_Fiyatlari_2017_2026.xlsx'

TR_HEADERS = {
'poz_no':'Poz No','unit':'Birim','latest_published_price':'Son Yayın Fiyatı','latest_period':'Son Yayın Dönemi','validity_date':'Geçerlilik Tarihi','source_url':'Kaynak','source_page':'Kaynak Sayfa','previous_period':'Önceki Yayın Dönemi','previous_price':'Önceki Fiyat','publication_count':'Yayın Sayısı','old_poz_numbers':'Eski Poz Numaraları','new_poz_no':'Yeni Poz Numarası','number_status':'Numara Durumu','mapping_confidence':'Eşleştirme Güveni','control_note':'Kontrol Notu',
'price':'Fiyat','period':'Dönem','raw_excerpt':'Ham Metin','mapping_period':'Eşleştirme Dönemi','old_poz_no':'Eski Poz No','new_poz_no':'Yeni Poz No','evidence':'Kanıt','mapping_type':'Eşleştirme Türü','kind':'Tür','status':'Durum','pages':'Sayfa Sayısı','mechanical_start_page':'Mekanik Başlangıç','mechanical_end_page':'Mekanik Bitiş','parsed_rows':'Alınan Satır','note':'Not'
}

def read_csv(path):
    if not path.exists(): return []
    with path.open(encoding='utf-8-sig', newline='') as f: return list(csv.DictReader(f))

def add_sheet(wb, title, rows, include_blank_name=False):
    ws=wb.create_sheet(title)
    keys=list(rows[0].keys()) if rows else []
    if include_blank_name:
        keys=[keys[0], 'poz_name_blank', *keys[1:]] if keys else ['poz_no','poz_name_blank']
    headers=[TR_HEADERS.get(k, 'Poz Tanımı' if k=='poz_name_blank' else k) for k in keys]
    ws.append(headers)
    for r in rows:
        vals=[]
        for k in keys:
            if k=='poz_name_blank': vals.append('')
            else:
                v=r.get(k,'')
                if k in {'latest_published_price','previous_price','price'}:
                    try: v=float(str(v).replace(',','.')) if v!='' else ''
                    except: pass
                elif k in {'source_page','publication_count','pages','mechanical_start_page','mechanical_end_page','parsed_rows'}:
                    try: v=int(v) if v!='' else ''
                    except: pass
                vals.append(v)
        ws.append(vals)
    fill=PatternFill('solid',fgColor='17365D'); font=Font(color='FFFFFF',bold=True)
    for c in ws[1]: c.fill=fill; c.font=font; c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; ws.row_dimensions[1].height=32
    for col in ws.columns:
        letter=col[0].column_letter
        maxlen=max(len(str(c.value or '')) for c in col[:200])
        ws.column_dimensions[letter].width=min(max(maxlen+2,11),55)
        for c in col[1:]: c.alignment=Alignment(vertical='top',wrap_text=True)
    for h in ('Son Yayın Fiyatı','Önceki Fiyat','Fiyat'):
        if h in headers:
            idx=headers.index(h)+1
            for row in ws.iter_rows(min_row=2,min_col=idx,max_col=idx): row[0].number_format='#,##0.00'
    if ws.max_row>1 and ws.max_column>0:
        name='T'+''.join(ch for ch in title if ch.isalnum())[:20]
        tab=Table(displayName=name,ref=ws.dimensions); tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True)
        ws.add_table(tab)
    return ws

def main():
    master,raw,maps,sources=map(read_csv,[MASTER,RAW,MAP,SOURCES])
    if not master: raise SystemExit('Ana CSV bulunamadı veya boş.')
    wb=Workbook(); wb.remove(wb.active)
    info=wb.create_sheet('Açıklama')
    info.append(['ÇŞİDB / Yüksek Fen Kurulu Mekanik Pozları – Son Yayın Fiyatları'])
    info.append(['Kapsam','2017–2026 resmî İnşaat ve Tesisat Birim Fiyatları; ara dönem yayınları dâhil.'])
    info.append(['Yöntem','En yeni yayından eskiye gidilerek her pozun son yayımlanan fiyatı seçildi.'])
    info.append(['Poz numarası değişiklikleri','Yalnız resmî değiştirilen poz numaraları belgeleri kullanıldı.'])
    info.append(['Poz tanımları','İkinci işlemde doldurulmak üzere boş bırakıldı.'])
    info.append(['Oluşturulma',datetime.now(timezone.utc).isoformat()])
    info.merge_cells('A1:B1'); info['A1'].font=Font(size=16,bold=True,color='17365D'); info.column_dimensions['A'].width=34; info.column_dimensions['B'].width=110
    for row in info.iter_rows():
        for c in row: c.alignment=Alignment(vertical='top',wrap_text=True)
    add_sheet(wb,'Son Yayın Listesi',master,True)
    add_sheet(wb,'Tüm Yayın Kayıtları',raw)
    add_sheet(wb,'Eski Yeni Poz Eşleştirme',maps)
    add_sheet(wb,'Kaynak Kontrolü',sources)
    wb.save(XLSX)
    print(XLSX)

if __name__=='__main__': main()
